# See PyCharm help at https://www.jetbrains.com/help/pycharm/
# from selenium import webdriver
# import time
import pandas

# PRG1:ライブラリ設定
import sys
import os
import time
import shutil
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import io
from PIL import Image
import openpyxl as px
from openpyxl.styles import Alignment

def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    print_hi('PyCharm')

def main():
    # PRG2:クロール設定
    BASE_URL = 'https://jp.mercari.com'
    CUR_DIR = os.getcwd()
    args = sys.argv
    KEYWORD = 'ILCE-9'
    # for i in range(2, len(args)):
    #     KEYWORD += '+' + args[i]

    # 検索用URL生成
    URL_INI = BASE_URL + '/search/?keyword=' + KEYWORD + '&status=on_sale'
    url = URL_INI
    page_num = 1

    # 検索結果格納リスト・画像保存フォルダ作成
    result = []
    if os.path.isdir('./img') == False:
        os.mkdir('./img')

    # PRG3:スクレイピング実行
    while True:
        try:
            # ChromeでURLに接続
            options = Options()
            options.add_argument('--headless')
            # browser = webdriver.Chrome(executable_path='C:\webDriver\chromedriver.exe', options=options)
            browser = webdriver.Chrome(executable_path='C:\webDriver\chromedriver.exe')
            browser.get(url)
            time.sleep(5)
            html = browser.page_source.encode('utf-8')

            # Beautifulsoupで要素取得
            soup = BeautifulSoup(html, "html.parser")
            items_list = soup.find_all("li", attrs={'data-testid': 'item-cell'})

            # サムネイル画像取得
            for i, item in enumerate(items_list):
                time.sleep(1)
                item_title = item.find("mer-item-thumbnail").get('item-name')
                img_src = item.find("mer-item-thumbnail").get('src')
                response = requests.get(img_src)
                img_fname = img_src.split('?')[0].split('/')[-1]
                img_bin = io.BytesIO(response.content)
                pil_img = Image.open(img_bin)
                img_resize = scale_to_width(pil_img, 200)
                img_resize.save('./img/' + img_fname)

                # 取得結果をリストに保存
                result.append([item_title,
                               int(item.find("mer-item-thumbnail").get('price')),
                               os.path.join(CUR_DIR, 'img', img_fname),
                               BASE_URL + item.find('a').get('href')])

            # 次ページボタン処理
            next_button = soup.find('mer-button', attrs={'data-testid': "pagination-next-button"})
            if next_button:
                page_num += 1
                param = '&page_token=v1%3A' + str(page_num)
                next_url = URL_INI + param
                url = next_url
                next_url = ''
            else:
                break

        # エラー発生時の例外処理
        except Exception as e:
            message = "[エラー]" + "type:{0}".format(type(e)) + "\n" + "args:{0}".format(e.args)
            print(message)

        # Chrome終了処理
        finally:
            browser.close()
            browser.quit()

            # PRG4:Excelに結果出力
    wb = px.Workbook()
    ws = wb.active

    # 書式設定
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 40
    my_alignment = Alignment(vertical='top', wrap_text=True)

    # Excelヘッダー出力
    headers = ['アイテム名', '価格', 'サムネイル', 'URL']
    for i, header in enumerate(headers):
        ws.cell(row=1, column=i + 1, value=header)

    # 取得結果書き込み
    for y, row in enumerate(result):
        ws.row_dimensions[y + 2].height = 160
        for x, cell in enumerate(row):
            if x == 2:
                img = px.drawing.image.Image(cell)
                img.anchor = ws.cell(row=y + 2, column=x + 1).coordinate
                ws.add_image(img)
            elif x == 3:
                ws.cell(row=y + 2, column=x + 1).hyperlink = cell
                ws.cell(row=y + 2, column=x + 1).alignment = my_alignment
            else:
                ws.cell(row=y + 2, column=x + 1).value = cell
                ws.cell(row=y + 2, column=x + 1).alignment = my_alignment

    # Excelファイル保存
    xlname = './mercari_' + KEYWORD + '.xlsx'
    wb.save(xlname)

    # 保存した画像の削除
    shutil.rmtree('./img')

    print('--- END ---')


# PRG8:アスペクト比固定して画像リサイズ
def scale_to_width(img, width):
    height = round(img.height * width / img.width)
    return img.resize((width, height))


# スクリプトとして実行された場合の処理
if __name__ == '__main__':
    main()

# キーワード入力
# search_word = input("カメラ＝")

# # メルカリ
# url = 'https://www.mercari.com/jp/search/?keyword=カメラ'# + search_word
#
# # chromedriverの設定とキーワード検索実行
# #driver = webdriver.Chrome()
# driver = webdriver.Chrome(executable_path="C:\webDriver\chromedriver")
# driver.get(url)
#
# # ページカウントとアイテムカウント用変数
# page = 1
# item_num = 0
# item_urls = []
#
# while True:
#     print("Getting the page {} ...".format(page))
#     time.sleep(1)
#     items = driver.find_elements_by_class_name("items-box")
#     for item in items:
#         item_num += 1
#         item_url = item.find_element_by_css_selector("a").get_attribute("href")
#         print("item{0} url:{1}".format(item_num, item_url))
#         item_urls.append(item_url)
#     page += 1
#
#     try:
#         next_page = driver.find_element_by_css_selector("li.pager-next .pager-cell:nth-child(1) a").get_attribute(
#             "href")
#         driver.get(next_page)
#         print("next url:{}".format(next_page))
#         print("Moving to the next page...")
#     except:
#         print("Last page!")
#         break
#
# # アイテムカウントリセットとデータフレームセット
# item_num = 0
# columns = ["item_name", "cat1", "cat2", "cat3", "brand_name", "product_state", "price", "url"]
# df = pandas.DataFrame(columns=columns)
#
# try:  # エラーで途中終了時をtry～exceptで対応
#     # 取得した全URLを回す
#     for product_url in item_urls:
#         item_num += 1
#         print("Moving to the item {}...".format(item_num))
#         time.sleep(1)
#         driver.get(product_url)
#
#         item_name = driver.find_element_by_css_selector("h1.item-name").text
#         print("Getting the information of {}...".format(item_name))
#
#         cat1 = driver.find_element_by_css_selector(
#             "table.item-detail-table tbody tr:nth-child(2) td a:nth-child(1) div").text
#         cat2 = driver.find_element_by_css_selector(
#             "table.item-detail-table tbody tr:nth-child(2) td a:nth-child(2) div").text
#         cat3 = driver.find_element_by_css_selector(
#             "table.item-detail-table tbody tr:nth-child(2) td a:nth-child(3) div").text
#         try:  # 存在しない⇒a, divタグがない場合をtry～exceptで対応
#             brand_name = driver.find_element_by_css_selector(
#                 "table.item-detail-table tbody tr:nth-child(3) td a div").text
#         except:
#             brand_name = ""
#
#         product_state = driver.find_element_by_css_selector("table.item-detail-table tbody tr:nth-child(4) td").text
#         price = driver.find_element_by_xpath("//div[1]/section/div[2]/span[1]").text
#         price = price.replace("¥", "").replace(" ", "").replace(",", "")
#
#         print(cat1)
#         print(cat2)
#         print(cat3)
#         print(brand_name)
#         print(product_state)
#         print(price)
#         print(product_url)
#
#         se = pandas.Series([item_name, cat1, cat2, cat3, brand_name, product_state, price, product_url], columns)
#         df = df.append(se, ignore_index=True)
#         print("Item {} added!".format(item_num))
#
# except:
#     print("Error occurred! Process cancelled but the added items will be exported to .csv")
#
# df.to_csv("{}.csv".format('カメラ'), index=False, encoding="utf_8")
# driver.quit()
# print("Scraping is complete!")



# メインプログラム